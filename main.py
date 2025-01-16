"""

Daily KinderTales Audit Assistant

Enter date of audit and name of teacher
Enter name of each child 
Create a table to note completion of KinderTales requirements
Save/Export table as .xlsx or PDF file for records

"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


# Functions
def add_child_button_action():
    input_child = add_child_entry.get().strip()
    if input_child:
        row_id = table.insert("", "end", values=(input_child, "", ""))

        for col_id in range(1, 7):
            add_dropdown(row_id, col_id)

        add_child_entry.delete(0, "end")


def add_dropdown(row_id, col_id):
    # Get the bounding box for the specific cell
    bbox = table.bbox(row_id, col_id)
    if bbox:
        x, y, width, height = bbox

        # Create the dropdown and place it dynamically
        dropdown = ttk.Combobox(main_window, values=["Yes", "No"], state="readonly")
        dropdown.place(
            x=x + table.winfo_rootx() - main_window.winfo_rootx(),
            y=y + table.winfo_rooty() - main_window.winfo_rooty(),
            width=width,
            height=height,
        )

        # Preselect "No" as the default value
        dropdown.set("No")
        table.set(row_id, col_id, "No")  # Update cell value

        # Event to update table and remove dropdown
        def on_dropdown_select(event):
            table.set(row_id, col_id, dropdown.get())  # Update cell value
            print(f"Row {row_id} updated: {table.item(row_id)['values']}")  # Debugging

        # Bind dropdown events
        dropdown.bind("<FocusOut>", on_dropdown_select)
        dropdown.bind("<<ComboboxSelected>>", on_dropdown_select)


def calculate_grade():
    total_cells = 0
    yes_count = 0

    for row_id in table.get_children():
        row_values = table.item(row_id)["values"]
        print(f"{row_id}: {row_values}")

        for col_id in range(1, len(row_values)):
            cell_value = row_values[col_id]
            print(f"Cell [{row_id}, {col_id}] = {cell_value}")

            if cell_value == "Yes":
                yes_count += 1
            if cell_value in ["Yes", "No"]:
                total_cells += 1

    print(f"Yes Count: {yes_count}, Total Cells: {total_cells}")

    if total_cells > 0:
        grade = (yes_count / total_cells) * 100
    else:
        grade = 0

    result_label.configure(text=f"Grade: {grade:.2f}%")


def export_to_excel(save_path):
    data = []
    for row_id in table.get_children():
        data.append(table.item(row_id)["values"])

    df = pd.DataFrame(data, columns=columns)

    audit_date = date_of_audit_entry.get().strip()
    teacher_name = teacher_name_entry.get().strip()
    # metadata = f"Date of Audit: {audit_date}, \n Teacher: {teacher_name}"
    metadata = pd.DataFrame(
        [["Teacher's Name:", teacher_name], ["Date of Audit:", audit_date]],
        columns=["", ""],
    )

    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        metadata.to_excel(
            writer, index=False, header=False, startrow=0, sheet_name="Audit Report"
        )
        df.to_excel(
            writer, index=False, header=True, startrow=3, sheet_name="Audit Report"
        )

        # AutoFit columns
        sheet = writer.sheets["Audit Report"]
        
        for col in sheet.iter_cols():
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col[0].column_letter].width = adjusted_width


    print(f"Excel file saved at {save_path}")


def export_to_pdf(save_path):
    # Create PDF instance
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    # Title
    pdf.set_font("Arial", style="B", size=16)
    pdf.cell(280, 10, "KinderTales Audit Report", ln=True, align="C")

    # Add metadata
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, txt=f"Date of Audit: {date_of_audit_entry.get().strip()}", ln=True)
    pdf.cell(0, 10, txt=f"Teacher: {teacher_name_entry.get().strip()}", ln=True)
    pdf.ln(10)

    col_widths = [max(len(col), len(str(col))) * 5 for col in columns]
    total_width = sum(col_widths)

    scale_factor = 280 / total_width
    col_widths = [width * scale_factor for width in col_widths]

    pdf.set_font("Arial", style="B", size=12)
    for i, col in enumerate(columns):
        pdf.cell(col_widths[i], 10, col, border=1, align="C")
    pdf.ln()

    # Add table data
    pdf.set_font("Arial", size=10)
    for row_id in table.get_children():
        row_values = table.item(row_id)["values"]
        for i, value in enumerate(row_values):
            pdf.cell(col_widths[i], 10, str(value), border=1, align="C")
        pdf.ln()

    pdf.output(save_path)


def export_data_dialog():
    # Open dialog box to choose file type and location
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=(("Excel files", "*.xlsx"), ("PDF files", "*.pdf")),
        title="Save Audit Report",
    )

    if not file_path:
        return

    if file_path.endswith(".xlsx"):
        export_to_excel(file_path)
        ctk.CTkLabel(main_window, text=f"Exported to {file_path}").pack()
    elif file_path.endswith(".pdf"):
        export_to_pdf(file_path)
        ctk.CTkLabel(main_window, text=f"Exported to {file_path}").pack()


# CTk settings
ctk.set_appearance_mode("system")
ctk.set_default_color_theme("green")

# Main Window Setup
main_window = ctk.CTk()
main_window.title("KinderTales Audit Assistant")
main_window.geometry("1000x800")

# Frame for buttons
button_frame = ctk.CTkFrame(main_window)
button_frame.pack(pady=20)

# INPUT: Date of Audit, Teacher(s) Name, Add Child
date_of_audit_label = ctk.CTkLabel(button_frame, text="Date of Audit:")
date_of_audit_label.grid(row=0, column=0, pady=(10, 0), padx=(0, 10))
date_of_audit_entry = ctk.CTkEntry(button_frame)
date_of_audit_entry.grid(row=0, column=1, pady=(10, 0))

teacher_name_label = ctk.CTkLabel(button_frame, text="Teacher's Name:")
teacher_name_label.grid(row=1, column=0, pady=(10, 0), padx=(10, 10))
teacher_name_entry = ctk.CTkEntry(button_frame)
teacher_name_entry.grid(row=1, column=1, pady=(10, 0))

add_child_label = ctk.CTkLabel(button_frame, text="Child Name:")
add_child_label.grid(row=2, column=0, pady=(10, 0), padx=(10, 10))
add_child_entry = ctk.CTkEntry(button_frame)
add_child_entry.grid(row=2, column=1, pady=(10, 0))
add_child_button = ctk.CTkButton(
    button_frame, text="Add Child", command=add_child_button_action
)
add_child_button.grid(row=2, column=2, pady=(10, 0), padx=(10, 10))

# Frame for table
table_frame = ctk.CTkFrame(main_window)
table_frame.pack(pady=20, padx=20, fill="both", expand=True)

# Create Treeview Table
columns = (
    "Child",
    "Pictures (3)",
    "Daily Blurb",
    "Daily Updates",
    "Curriculum",
    "Name to Face",
    "Portfolio",
)
table = ttk.Treeview(table_frame, columns=columns, show="headings")

for col in columns:  # Add and format columns
    table.heading(col, text=col)
    table.column(col, anchor="center", width=100)

# Pack the table
table.pack(pady=20, padx=20, fill="both", expand=True)

# Calculate grade button
calculate_button = ctk.CTkButton(
    main_window, text="Calculate Grade", command=calculate_grade
)
calculate_button.pack(pady=(10, 0))
result_label = ctk.CTkLabel(main_window, text="Grade: 0.00%")
result_label.pack()

# Export button
save_button = ctk.CTkButton(main_window, text="Export")
save_button.configure(command=export_data_dialog)
save_button.pack(pady=15)

main_window.mainloop()
